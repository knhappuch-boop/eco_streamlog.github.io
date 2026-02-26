from flask import Flask, request, jsonify, send_file
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

app = Flask(__name__)

# Path to store customer records
EXCEL_FILE = "customer_records.xlsx"

def init_excel_file():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        
            # Create Quote sheet
            ws = wb.active
            ws.title = "Quotes"
        
            # Add quote headers
            quote_headers = ["Name", "Email", "Delivery Details", "Submission Date & Time"]
            ws.append(quote_headers)
        
            # Create Message sheet
            ws_msg = wb.create_sheet("Messages")
            msg_headers = ["Name", "Email", "Phone", "Message Type", "Message", "Submission Date & Time"]
            ws_msg.append(msg_headers)
        
            # Style header rows
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
            for cell in ws_msg[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 25
        
            ws_msg.column_dimensions['A'].width = 20
            ws_msg.column_dimensions['B'].width = 30
            ws_msg.column_dimensions['C'].width = 15
            ws_msg.column_dimensions['D'].width = 15
            ws_msg.column_dimensions['E'].width = 40
            ws_msg.column_dimensions['F'].width = 25
        
        wb.save(EXCEL_FILE)

@app.route('/submit-quote', methods=['POST'])
def submit_quote():
    """Handle quote form submissions and save to Excel"""
    try:
        data = request.json
        
        # Validate required fields
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        details = data.get('details', '').strip()
        
        if not name or not email or not details:
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        
        # Initialize Excel file if it doesn't exist
        init_excel_file()
        
        # Load workbook and add new row
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Quotes']
        
        submission_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([name, email, details, submission_time])
        
        # Save workbook
        wb.save(EXCEL_FILE)
        
        return jsonify({'success': True, 'message': 'Thank you for contacting Eco-Stream Logistics! Your inquiry has been recorded.'}), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/submit-message', methods=['POST'])
def submit_message():
    """Handle message form submissions and save to Excel"""
    try:
        data = request.json
        
        # Validate required fields
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        msg_type = data.get('type', '').strip()
        message = data.get('message', '').strip()
        phone = data.get('phone', '').strip()
        
        if not name or not email or not msg_type or not message:
            return jsonify({'success': False, 'message': 'Name, Email, Message Type, and Message are required'}), 400
        
        # Initialize Excel file if it doesn't exist
        init_excel_file()
        
        # Load workbook and add new row to Messages sheet
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb['Messages']
        
        submission_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([name, email, phone, msg_type, message, submission_time])
        
        # Save workbook
        wb.save(EXCEL_FILE)
        
        msg_type_display = msg_type.capitalize()
        return jsonify({'success': True, 'message': f'Thank you! Your {msg_type_display} has been recorded.'}), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'ok'}), 200


@app.route('/download-records', methods=['GET'])
def download_records():
    """Serve the customer_records.xlsx file for download"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'success': False, 'message': 'Records file not found.'}), 404
        # send as attachment so browser downloads the file
        return send_file(EXCEL_FILE, as_attachment=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

if __name__ == '__main__':
    # Initialize Excel file on startup
    init_excel_file()
    # Run the Flask app on localhost:5000
    app.run(debug=True, port=5000)
